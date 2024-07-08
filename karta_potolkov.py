
from curl_cffi import requests
from openpyxl import Workbook, load_workbook

session = requests.Session()
headers = {
    'Host': 'xn--80aag3acfuacfjyd.xn--p1ai',
    'Accept': 'application/json, text/plain, */*',
    # 'Accept-Encoding': 'gzip, deflate',
    'User-Agent': 'okhttp/4.9.2',
}
session.headers.update(headers)
flattens = []

ids = []

for page in range(1, 57+1):

    params = {
        'map_lon_max': '168.1245855364972',
        'map_lon_min': '3.105644307287889',
        'map_lat_max': '80.99435213439257',
        'map_lat_min': '-60.13880870465776',
        'user_lat': '55.7569618',
        'user_lon': '37.6150094',
        'page': page,
        'per_page': '15',
        'sorting': '-provider__verified,-rating',
        'manufacturer_ids': '',
        'material_ids': '',
    }

    response = session.get('https://xn--80aag3acfuacfjyd.xn--p1ai/api/warehouses/', params=params)

    print(response, page)
    resp = response.json().get('results')
    for r in resp:
        ids.append(r.get('id'))


for key, id in enumerate(ids, 1):
    response = session.get(f'https://xn--80aag3acfuacfjyd.xn--p1ai/api/warehouses/{id}?user_lat=55.7569624&user_lon=37.6150095')
    print(response, key, len(ids))
    resp = response.json()
    phones = ' '.join(resp.get('phones', []))
    company = resp.get('provider', {}).get('name', '')
    gorod = resp.get('address', {}).get('city', '')


    flattens.append([company, gorod, phones])

column_xlsx = ['Название компании', 'Город', 'Номер телефона']

file_path_write = 'output.xlsx'
workbook = Workbook()
worksheet = workbook.active
worksheet.title = "Data"
worksheet.append(column_xlsx)
workbook.save(file_path_write)

workbook = load_workbook(file_path_write)
worksheet = workbook.active
for rrik in flattens:
    worksheet.append(rrik)
workbook.save(file_path_write)
