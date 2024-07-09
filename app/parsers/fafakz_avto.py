
from bs4 import BeautifulSoup
from curl_cffi import requests
from openpyxl import Workbook, load_workbook
import re
import time

session = requests.Session()

vehicle_types = {
    "Крытая": -1,
    "Открытая": -2,
    "Любая": 0,
    "Тент": 1,
    "Рефрижератор": 2,
    "Изотерм": 3,
    "Бортовой": 4,
    "Автобус": 5,
    "Автовоз": 6,
    "Автовышка": 7,
    "Автотранспортер": 8,
    "Балковоз": 9,
    "Бензовоз": 10,
    "Бетоновоз": 11,
    "Битумовоз": 12,
    "Газовоз": 13,
    "Зерновоз": 14,
    "Конт.площадка": 15,
    "Кормовоз": 16,
    "Кран": 17,
    "Лесовоз": 18,
    "Манипулятор": 19,
    "Микроавтобус": 20,
    "Муковоз": 21,
    "Открытый Конт.": 22,
    "Панелевоз": 23,
    "Погрузчик": 24,
    "Самосвал": 25,
    "Седельный Тягач": 26,
    "Скотовоз": 27,
    "Трал": 29,
    "Трубовоз": 31,
    "Цельнометалл.": 32,
    "Цементовоз": 33,
    "Цистерна": 34,
    "Шаланда": 35,
    "Щеповоз": 36,
    "Эвакуатор": 37,
    "Экскаватор": 38
}


punkt_a = 'Казахстан'

punkt_b = 'Россия'

def find_vehicle_type(cargo_info):
    for vehicle in vehicle_types:
        if re.search(r'\b' + vehicle[:4], cargo_info, re.IGNORECASE):
            return vehicle
    return "Неизвестный тип"

vehicl_typ = vehicle_types.get(input('Введите тип машины: ').capitalize(), '0')

cookies = {
            '_ym_uid': '1719577624223202900',
            '_ym_d': '1719577624',
            '_ga_3E9ZB9DLH7': 'GS1.1.1719922003.3.1.1719923330.0.0.0',
            '_ga': 'GA1.1.395825761.1719577624',
            '_ga_N8G7P3XR38': 'GS1.1.1719922003.3.1.1719923330.0.0.0',
            'mid': '10384507',
            'mid2': 'dd41475f3aea8698994d8fc04081d0ed',
            '_gid': 'GA1.2.778041633.1719918404',
            '_ym_isad': '1',
            'bbuserid': '239374',
            'bbpassword': '1eeea5178d88f76a0d3fd235cfbb77ca',
            'c_uid': '239374',
            'puid': '58a0086107c9b676c5bece0c1eda0dce',
            'PHPSESSID': 'jdo1f48ttq5o56bafc05t61r97',
            'bbuserid': '239374',
            'bbpassword': '1eeea5178d88f76a0d3fd235cfbb77ca',
            'ref_icon': '0',
            '_ga_295751JDV1': 'GS1.1.1719923347.1.1.1719923466.60.0.0',
            '_gat_gtag_UA_113967512_1': '1',
            'whatsapp_drugu': '1',
        }

headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:127.0) Gecko/20100101 Firefox/127.0',
    'Accept': 'text/html, */*',
    'Accept-Language': 'ru-RU,ru;q=0.8,en-US;q=0.5,en;q=0.3',
    # 'Accept-Encoding': 'gzip, deflate, br, zstd',
    'X-Requested-With': 'XMLHttpRequest',
    'Alt-Used': 'fa-fa.kz',
    'Connection': 'keep-alive',
    'Referer': 'https://fa-fa.kz/search_load/',
    # 'Cookie': '_ga_3E9ZB9DLH7=GS1.1.1719954953.12.1.1719954955.0.0.0; _ga=GA1.2.1858377429.1717621778; _ga_N8G7P3XR38=GS1.1.1719954953.13.1.1719954955.0.0.0; _ym_uid=1717621778108757811; _ym_d=1717621778; mid=10394511; mid2=7a001a6f5229051f4ada1810bdff3668; c_uid=239374; _ga_295751JDV1=GS1.1.1719902909.6.1.1719904785.60.0.0; _gid=GA1.2.2034431614.1719829465; _ym_isad=1; bbuserid=239374; bbpassword=c97a60f50e3466e948d29aee676caca3; puid=58a0086107c9b676c5bece0c1eda0dce; PHPSESSID=98rpv1qcgahrriqckbgshp27t6; _gat_gtag_UA_76271908_1=1',
    'Sec-Fetch-Dest': 'empty',
    'Sec-Fetch-Mode': 'cors',
    'Sec-Fetch-Site': 'same-origin',
    # Requests doesn't support trailers
    # 'TE': 'trailers',
}

params = {
    'n': '1',
    'str': punkt_a,
}

response = session.get('https://fa-fa.kz/city.php', params=params,headers=headers)
print(response)
soup = BeautifulSoup(response.text, 'html.parser')
punkt_a = soup.find('div').text

params = {
    'n': '1',
    'str': punkt_b,
}

response = session.get('https://fa-fa.kz/city.php', params=params,headers=headers)
soup = BeautifulSoup(response.text, 'html.parser')
punkt_b = soup.find('div').text

params = {
    'blank': '1',
}

headers = {
    'Host': 'fa-fa.kz',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:127.0) Gecko/20100101 Firefox/127.0',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8',
    'Accept-Language': 'ru-RU,ru;q=0.8,en-US;q=0.5,en;q=0.3',
    # 'Accept-Encoding': 'gzip, deflate',
    'Content-Type': 'application/x-www-form-urlencoded',
    # 'Content-Length': '201',
    'Origin': 'https://fa-fa.kz',
    'Referer': 'https://fa-fa.kz/search_car/',
    'Upgrade-Insecure-Requests': '1',
    'Sec-Fetch-Dest': 'document',
    'Sec-Fetch-Mode': 'navigate',
    'Sec-Fetch-Site': 'same-origin',
    'Sec-Fetch-User': '?1',
    'Priority': 'u=1',
    # Requests doesn't support trailers
    # 'Te': 'trailers',
    # 'Cookie': '_ga_3E9ZB9DLH7=GS1.1.1720046845.14.1.1720047817.0.0.0; _ga=GA1.2.1858377429.1717621778; _ga_N8G7P3XR38=GS1.1.1720046845.15.1.1720047817.0.0.0; _ym_uid=1717621778108757811; _ym_d=1717621778; mid=10394511; mid2=7a001a6f5229051f4ada1810bdff3668; c_uid=239374; _ga_295751JDV1=GS1.1.1719902909.6.1.1719904785.60.0.0; bbuserid=239374; bbpassword=c97a60f50e3466e948d29aee676caca3; puid=58a0086107c9b676c5bece0c1eda0dce; PHPSESSID=236ls96rb0cu5hr2nu0a5hqba1; _gid=GA1.2.1138017272.1720043543; _ym_isad=1',
}
data = f'type_add=car&form_type_kuz=&form_type=&type21={vehicl_typ}&adr=0&w_ot=&w_do=&v_ot=&v_do=&dat1=0&dat2=7&siz1=&siz2=&siz3=&sload=0&city1={punkt_a}&rad1=&city4={punkt_b}&rad4=&car_search='
response = session.post('https://fa-fa.kz/search_car/', params=params,  headers=headers, data=data)
print(response)
if response.status_code == 302:
    sid = response.headers.get('Location', '').split('sid=')[-1]

soup = BeautifulSoup(response.text, 'html.parser')
td_tag = soup.find('td', class_='search_s2')

if td_tag:
    onclick_value = td_tag.get('onclick')
    if onclick_value:
        # Разбиваем строку по символу '=' и берем последний элемент
        sid = onclick_value.split('=')[-1].split('&')[0].split("'")[0]
        print("SID:", sid)
    else:
        print("Атрибут onClick не найден.")
else:
    sid = ''

params = {
    'sid': sid
}




flattens = []

pattern = re.compile(r'^res_\d+$')

for page in range(1, +1):
    response = session.get(f'https://fa-fa.kz/search_car/{page}/', params=params, headers=headers)
    print(response)
    soup = BeautifulSoup(response.text, 'html.parser')
    for div in soup.find_all('div', id=pattern):
        route = div.find('font', class_='o_dest').text.strip()

        id_div = div['id'].split('_')[-1]

        print(route)
        departure_country_code = route.split('-')[0].split('(')[-1].strip(')').strip('(')
        departure_city = route.split('(')[0].strip()

        arrival_city = route.split('-')[-1].strip()
        if arrival_city == 'Любое направление':
            arrival_city = arrival_country_code = 'Любое направление'
        else:
            arrival_city = route.split('-')[-1].split('(')[0].strip()

            arrival_country_code = route.split('-')[-1].split('(')[-1].strip(')')

        dates = div.find_all('b')[0:2]
        date_from = dates[0].text.strip()
        date_to = dates[1].text.strip()
        date = f"{date_from} - {date_to}"

        # Извлечение информации о типе машины/груза
        cargo_info = div.find_all('td', valign='top')[1].find(text=True, recursive=False)
        print(cargo_info)

        # Поиск типа машины/груза
        cargo_type = find_vehicle_type(cargo_info)

        # Извлечение веса
        weight_match = div.find('td', attrs={'style': 'width:8%;max-width:80px;'}).find('b').text.split('/')[0].strip()
        print(weight_match)

        price_info = div.find_all('td', valign='top')[2].get_text(separator=' ').strip()


        price_match = re.search(r'(\d[\d\s\.\,]*\d)', price_info)
        price = price_match.group(0) if price_match and len(
            price_match.group(0).replace(' ', '').replace(',', '')) >= 4 else "Неизвестная цена"

        params = {
            'bid': id_div,
        }

        response = session.get('https://fa-fa.kz/index/car_contacts/', params=params, cookies=cookies,
                                headers=headers)
        print(f'contact {response}')
        soup = BeautifulSoup(response.text, 'html.parser')
        contact = ''
        for i in soup.find_all('a'):
            str1 = i.text.strip()
            contact += f'{str1} '

        time.sleep(1)

        flattens.append([route, departure_country_code, departure_city, arrival_city, arrival_country_code, date, cargo_type, weight_match, price, contact , 'fafa-kz'])

column_xlsx = ["Маршрут", "Город отправления",
        "Страна отправления",
        "Город прибытия",
        "Страна прибытия", 'Дата', 'Тип машины/груза', 'Вес', 'Цена', 'Контакты' , 'Сайт']
file_path_write = 'output.xlsx'
workbook = Workbook()
worksheet = workbook.active
worksheet.title = "Data"
worksheet.append(column_xlsx)
workbook.save(file_path_write)

workbook = load_workbook(file_path_write)
worksheet = workbook.active
for rrik in flattens:
    worksheet.append(rrik)
workbook.save(file_path_write)






