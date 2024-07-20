import requests
from datetime import datetime, timedelta
from itertools import combinations
from openpyxl import Workbook, load_workbook

transport_type = {
            "контейнер": "2",
            "тентованный": "1",
            "рефрижератор": "4",
            "реф. с перегородкой": "281474976710656",
            "реф. мультирежимный": "562949953421312",
            "изотермический": "8",
            "фургон": "16",
            "микроавтобус": "32",
            "цельнометалл.": "64",
            "бортовой": "128",
            "открытый конт.": "1024",
            "пикап": "68719476736",
            "самосвал": "4096",
            "кормовоз": "4194304",
            "коневоз": "137438953472",
            "контейнеровоз": "2097152",
            "манипулятор": "256",
            "площадка без бортов": "70368744177664",
            "шаланда": "8192",
            "негабарит": "18726594281984",
            "кран": "8388608",
            "автотранспортер": "131072",
            "автоцистерна": "8589934592",
            "лесовоз": "16777216",
            "панелевоз": "2048",
            "ломовоз": "1125899906842624",
            "трубовоз": "1073741824",
            "седельный тягач": "67108864",
            "низкорамный": "512",
            "низкорам.платф.": "34359738368",
            "телескопический": "1099511627776",
            "газовоз": "524288",
            "трал": "536870912",
            "автобус": "16384",
            "скотовоз": "134217728",
            "стекловоз": "268435456",
            "все открытые": "70368744191104",
            "цементовоз": "2147483648",
            "щеповоз": "4294967296",
            "муковоз": "33554432",
            "автовоз": "32768",
            "автовышка": "65536",
            "бетоновоз": "262144",
            "битумовоз": "2199023255552",
            "балковоз(негабарит)": "17592186044416",
            "эвакуатор": "17179869184",
            "бензовоз": "274877906944",
            "вездеход": "549755813888",
            "реф.-тушевоз": "4398046511104",
            "пирамида": "8796093022208",
            "пухтовоз": "140737488355328",
            "рулоновоз": "35184372088832",
            "все закр.+изотерм": "91",
            "зерновоз": "1048576",
            "реф.+изотерм": "844424930131980",
            "грузопассажирский": "2251799813685248",
            "клюшковоз": "4503599627370496",
            "мусоровоз": "9007199254740992",
            "jumbo": "18014398509481984",
            "20' танк-контейнер": "36028797018963968",
            "40' танк-контейнер": "72057594037927936",
            "мега фура": "144115188075855872",
            "допельшток": "288230376151711744",
            "Раздвижной полуприцеп 20'/40'": "576460752303423488"
        }

cargo_type = {'автошины': 1, 'алкогольные напитки': 2, 'безалкогольные напитки': 3, 'бумага': 4, 'бытовая техника': 5, 'грибы': 6, 'древесина': 7, 'древесный уголь': 8, 'зерно и семена (в упаковке)': 9, 'изделия из кожи': 10, 'изделия из металла': 11, 'канц. товары': 13, 'ковры': 14, 'компьютеры': 15, 'консервы': 16, "40' контейнер": 17, 'макулатура': 18, 'мебель': 19, 'медикаменты': 20, 'металл': 21, 'металлолом': 22, 'молоко сухое': 23, 'мороженое': 24, 'мясо': 25, 'нефтепродукты': 26, 'оборудование и запчасти': 27, 'обувь': 28, 'овощи': 29, 'одежда': 30, 'парфюмерия и косметика': 31, 'пиво': 32, 'пластик': 33, 'продукты питания': 34, 'птица ': 35, 'изделия из резины': 36, 'рыба (неживая)': 37, 'сантехника': 38, 'сахар': 39, 'сборный груз': 40, 'стекло и фарфор': 41, 'стройматериалы': 42, 'табачные изделия': 43, 'тара и упаковка': 44, 'текстиль': 45, 'тнп': 46, 'торф': 47, 'транспортные средства': 50, 'удобрения': 51, 'фрукты': 52, 'хим. продукты опасные': 53, 'хим. продукты неопасные': 54, 'хозтовары': 55, 'шкуры мокросоленые': 56, 'электроника': 57, 'ягоды': 58, 'другой': 59, 'дсп': 60, 'утеплитель': 61, 'кирпич': 62, 'трубы': 63, 'лдсп': 64, 'фанера': 65, 'минвата': 66, 'пенопласт': 67, 'гофрокартон': 68, 'напитки': 69, 'стеклотара (бутылки и др.)': 70, 'мука': 71, 'поддоны': 73, 'чипсы': 74, 'соки': 75, 'цемент': 76, 'кондитерские изделия': 77, 'кабель': 78, 'холодильное оборудование': 79, 'доски': 80, 'пиломатериалы': 81, 'бытовая химия': 82, 'двп': 83, "20' контейнер": 84, 'крупа': 85, 'металлопрокат': 86, 'вагонка': 87, 'ферросплавы': 88, 'кормовые/пищевые добавки': 89, 'игрушки': 90, 'оборудование медицинское': 91, 'зерно и семена (насыпью)': 92, 'цветы': 93, 'шпалы': 94, 'жби': 95, 'гипс': 96, 'газосиликатные блоки': 97, 'арматура': 98, 'сэндвич-панели': 100, 'двери': 101, 'домашний переезд': 102, 'огнеупорная продукция': 103, 'инструмент': 105, 'люди': 106, 'соль': 107, 'мел': 108, 'песок': 109, 'щебень': 110, 'автомобиль(ли)': 111, 'балки надрессорные': 112, 'ж/д запчасти (прочие)': 113, 'боковая рама': 114, 'профлист': 206, "20' танк-контейнер": 207, "40' танк-контейнер": 208, 'колесная пара': 209, 'сонк (кп)': 210, 'поглощающий аппарат': 211, "20' контейнер hc": 212, "40' контейнер hc": 213, "45' контейнер (стар.)": 214, "45' контейнер (нов.)": 215, "20' реф.контейнер": 216, "40' реф.контейнер": 217, "45' реф.контейнер": 218, 'животные': 219, "40' реф.контейнер hc": 220}

def get_transport_id_by_initials(initials, transport_dict):
    # Приведем первые 3 буквы к нижнему регистру для сравнения
    initials = initials[:3].lower()
    for transport, id in transport_dict.items():
        if transport[:3].lower() == initials:
            return id
    return None



class AtiSu:

    def __init__(self, punkt_a, punkt_b, vehicl_type, cargo_name):
        self.punkt_a = punkt_a
        self.punkt_b = punkt_b
        self.vehicle_type = 0
        if vehicl_type:
            vehicle_type = vehicl_type.split(' ')
            for vh in vehicle_type:
                self.vehicle_type += get_transport_id_by_initials(vh, transport_type)


        self.cargo_name = []
        if cargo_name:
            cargo_name = cargo_name.split(' ')
            for cn in cargo_name:
                self.cargo_name.append(cargo_type.get(cn.lower(), 1))

    def convert_date_to_short_format(self, date_str):
        # Исходная дата в формате ISO 8601
        original_date = datetime.fromisoformat(date_str.rstrip('Z'))

        # Форматируем дату в строковый формат DD.MM
        formatted_date = original_date.strftime('%d.%m')

        # Возвращаем отформатированную дату
        return formatted_date

    def find_combination(self, target_number, transport_type):
        # Преобразуем значения словаря в числа
        transport_type_values = {key: int(value) for key, value in transport_type.items()}

        # Перебираем все возможные комбинации ключей
        for r in range(1, len(transport_type_values) + 1):
            for keys_combination in combinations(transport_type_values.keys(), r):
                # Вычисляем сумму значений для текущей комбинации ключей
                current_sum = sum(transport_type_values[key] for key in keys_combination)

                # Если сумма равна заданному числу, возвращаем найденную комбинацию ключей
                if current_sum == target_number:
                    return keys_combination

        # Если не найдено подходящей комбинации
        return None

    def run_gruz(self):
        cookies = {
            'auth_visit': '1',
            '_ga_14VPSGD0HN': 'GS1.1.1719835902.1.1.1719836588.0.0.0',
            '_ymab_param': 'w6rHAKlXzdvSgU0nf2aRFae9LFByZ79GB_7LmXUPzYuuQQn2hkp2IbJc176NcfqmVNVUmtmfw0AfGizvhVpaa5qKwJM',
            'tmr_detect': '0%7C1719836202990',
            '_ga_Z6YM1FRK5D': 'GS1.2.1719835904.1.1.1719836379.60.0.0',
            '_dc_gtm_UA-224067-1': '1',
            'tmr_lvid': '089160966f073ff8ad61ae20d3cc4c53',
            '_ym_visorc': 'w',
            'sid': '7a93fa4ba51e45b1a4584178597c43d5',
            'tmr_lvidTS': '1719835903012',
            'atiapp': '1',
            'startpage': 'idatisulogin',
            '_ym_uid': '1719835903778768010',
            'uicult': 'ru-RU',
            'X-Ati-Android-Client': '234',
            'AccessToken': '293bb1743aef4b9cb4eba530ca69f76f',
            '_ym_isad': '2',
            'lastpage': 'idatisulogin',
            'uicult2': 'ru',
            '_ym_d': '1719835903',
            '_gid': 'GA1.2.59457443.1719835903',
            'domain_sid': 'XB1DyqkBgV6bWe41fxizf%3A1719835904377',
            '_ga': 'GA1.2.46640107.1719835903',
            '_gcl_au': '1.1.849499656.1719835901',
            'did': 'mnBMhODY96nusg3m1X0fVqUHhPNGfbeAWah%2BlfJ4sBk%3D',
        }

        headers = {
            'Host': 'api.ati.su',
            'Authorization': 'Bearer 293bb1743aef4b9cb4eba530ca69f76f',
            'Content-Type': 'application/json',
            'X-Ati-Android-Client': '234',
            'Manufacturer': 'samsung',
            'Model': 'SM-S906N',
            'Su.ati.android': '234',
            'Accept': 'application/json',
            # 'Accept-Encoding': 'gzip, deflate',
            'User-Agent': 'okhttp/4.12.0',
            # 'Cookie': 'auth_visit=1; _ga_14VPSGD0HN=GS1.1.1719835902.1.1.1719836588.0.0.0; _ymab_param=w6rHAKlXzdvSgU0nf2aRFae9LFByZ79GB_7LmXUPzYuuQQn2hkp2IbJc176NcfqmVNVUmtmfw0AfGizvhVpaa5qKwJM; tmr_detect=0%7C1719836202990; _ga_Z6YM1FRK5D=GS1.2.1719835904.1.1.1719836379.60.0.0; _dc_gtm_UA-224067-1=1; tmr_lvid=089160966f073ff8ad61ae20d3cc4c53; _ym_visorc=w; sid=7a93fa4ba51e45b1a4584178597c43d5; tmr_lvidTS=1719835903012; atiapp=1; startpage=idatisulogin; _ym_uid=1719835903778768010; uicult=ru-RU; X-Ati-Android-Client=234; AccessToken=293bb1743aef4b9cb4eba530ca69f76f; _ym_isad=2; lastpage=idatisulogin; uicult2=ru; _ym_d=1719835903; _gid=GA1.2.59457443.1719835903; domain_sid=XB1DyqkBgV6bWe41fxizf%3A1719835904377; _ga=GA1.2.46640107.1719835903; _gcl_au=1.1.849499656.1719835901; did=mnBMhODY96nusg3m1X0fVqUHhPNGfbeAWah%2BlfJ4sBk%3D',
        }

        session = requests.Session()

        session.headers.update(headers)
        session.cookies.update(cookies)


        headers = {
            'Content-Type': 'application/json',
        }


        json_data = {
            'country_id': 10,
            'prefix': self.punkt_a,
            'suggestion_types': 23,
        }

        response = requests.post('https://api.ati.su/gw/gis-dict/public/v1/autocomplete/suggestions', headers=headers, json=json_data)

        punkt_a = response.json().get('suggestions', [{}])[0].get('country', {}).get('id', '')


        json_data = {
            'country_id': 10,
            'prefix': self.punkt_b,
            'suggestion_types': 23,
        }

        response = requests.post('https://api.ati.su/gw/gis-dict/public/v1/autocomplete/suggestions', headers=headers, json=json_data)

        punkt_b = response.json().get('suggestions', [{}])[0].get('country', {}).get('id', '')



        json_data = {
              "exclude_disputed_territory": False,
              "filter": {
                "cargo_types": [

                ],
                "change_date": 0,
                "dates": {
                  "date_option": "today-plus"
                },
                "from": {
                  "exact_only": False,
                  "id": punkt_a,
                  "type": 0
                },
                "with_auction": False,
                "sorting_type": 2,
                "to": {
                  "exact_only": False,
                  "id": punkt_b,
                  "type": 0
                },
                "truck_type": 16,
                "with_dimensions": False
              },
              "items_per_page": 100,
              "page": 1
            }


        json_data['filter']['truck_type'] = self.vehicle_type
        json_data['filter']['cargo_types'].extend(self.cargo_name)

        print(json_data)

        response = session.post(
            'https://api.ati.su/mobile/v1.0/loads/search',
            cookies=cookies,
            headers=headers,
            json=json_data,
        )
        print(response)
        resp = response.json().get('total_count', 0)
        print(resp)
        flattens = []
        if resp:
            for page in range(1, int(int(int(resp)/100) + 2)):
                json_data['page'] = page
                response = session.post(
                    'https://api.ati.su/mobile/v1.0/loads/search',
                    cookies=cookies,
                    headers=headers,
                    json=json_data,
                )
                print(response)
                resp = response.json()

                loads = resp.get('loads', [])
                for load in loads:
                    gorod_a = str(load.get('loading', {}).get('city_id', 9434))
                    gorod_b = str(load.get('unloading', {}).get('city_id', 3058))
                    print(gorod_a, gorod_b)
                    cities = resp.get('cities', {})
                    name_gorod_a = cities.get(gorod_a, {}).get('full_name', '')
                    name_gorod_b = cities.get(gorod_b, {}).get('full_name', '')

                    countries = load.get('route', {}).get('country', 'RUS-KAZ')
                    country_a = countries.split('-')[0]
                    country_b = countries.split('-')[-1]

                    first_date = load.get('first_date', '')
                    last_date = load.get('last_date', '')

                    if first_date:
                        first_date = self.convert_date_to_short_format(first_date)
                    if last_date:
                        last_date = self.convert_date_to_short_format(last_date)

                    transport_number = load.get('transport', {}).get('car_type_bit_mask', 1)
                    type_tran = ' '.join(self.find_combination(transport_number, transport_type))

                    weight = load.get('cargo', {}).get('weight', '')

                    price = load.get('payment', {}).get('sum_without_nds', 'Неизвестная цена')

                    ati_id = load.get('ati_id', 1437455)
                    contact_id1 = load.get('contact_id1', '')
                    contact_id2 = load.get('contact_id2', '')
                    contact_id3 = load.get('contact_id3', '')
                    contact_id4 = load.get('contact_id4', '')

                    acc = resp.get('accounts', {}).get(ati_id, {}).get('contacts', [])
                    telephones = []

                    email = ''

                    for ac in acc:
                        if ac.get('id', 0) in [contact_id1, contact_id2, contact_id3, contact_id4]:

                            telephones.append(ac.get('telephone', '') if ac.get('telephone', '') else '')
                            telephones.append(ac.get('mobile', '') if ac.get('mobile', '') else '')
                            id_ac = ac.get('id', 0)
                            try:
                                response = session.get(
                                    f'https://api.ati.su/mobile/v1.2/firms/contact/{ati_id}/{id_ac}/email'
                                )
                                email = response.json().get('email', '')
                            except:
                                print(f'Не удалось получить емеил https://api.ati.su/mobile/v1.2/firms/contact/{ati_id}/{id_ac}/email')
                                email = ''

                    telephones = ';'.join(telephones)

                    flattens.append([name_gorod_a, country_a, name_gorod_b, country_b, f'{first_date}-{last_date}', type_tran, weight, price, f'{telephones} {email}', 'ati-su'])

        column_xlsx = ["Город отправления",
                       "Страна отправления",
                       "Город прибытия",
                       "Страна прибытия", 'Дата', 'Тип машины/груза', 'Вес', 'Цена', 'Контакты', 'Сайт']

        file_path_write = 'output.xlsx'
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Data"
        worksheet.append(column_xlsx)
        workbook.save(file_path_write)

        workbook = load_workbook(file_path_write)
        worksheet = workbook.active
        for rrik in flattens:
            worksheet.append(rrik)
        workbook.save(file_path_write)

    def run_avto(self):
        cookies = {
            'auth_visit': '1',
            '_ga_14VPSGD0HN': 'GS1.1.1719835902.1.1.1719836588.0.0.0',
            '_ymab_param': 'w6rHAKlXzdvSgU0nf2aRFae9LFByZ79GB_7LmXUPzYuuQQn2hkp2IbJc176NcfqmVNVUmtmfw0AfGizvhVpaa5qKwJM',
            'tmr_detect': '0%7C1719836202990',
            '_ga_Z6YM1FRK5D': 'GS1.2.1719835904.1.1.1719836379.60.0.0',
            '_dc_gtm_UA-224067-1': '1',
            'tmr_lvid': '089160966f073ff8ad61ae20d3cc4c53',
            '_ym_visorc': 'w',
            'sid': '7a93fa4ba51e45b1a4584178597c43d5',
            'tmr_lvidTS': '1719835903012',
            'atiapp': '1',
            'startpage': 'idatisulogin',
            '_ym_uid': '1719835903778768010',
            'uicult': 'ru-RU',
            'X-Ati-Android-Client': '234',
            'AccessToken': '293bb1743aef4b9cb4eba530ca69f76f',
            '_ym_isad': '2',
            'lastpage': 'idatisulogin',
            'uicult2': 'ru',
            '_ym_d': '1719835903',
            '_gid': 'GA1.2.59457443.1719835903',
            'domain_sid': 'XB1DyqkBgV6bWe41fxizf%3A1719835904377',
            '_ga': 'GA1.2.46640107.1719835903',
            '_gcl_au': '1.1.849499656.1719835901',
            'did': 'mnBMhODY96nusg3m1X0fVqUHhPNGfbeAWah%2BlfJ4sBk%3D',
        }

        headers = {
            'Host': 'api.ati.su',
            'Authorization': 'Bearer 293bb1743aef4b9cb4eba530ca69f76f',
            'Content-Type': 'application/json',
            'X-Ati-Android-Client': '234',
            'Manufacturer': 'samsung',
            'Model': 'SM-S906N',
            'Su.ati.android': '234',
            'Accept': 'application/json',
            # 'Accept-Encoding': 'gzip, deflate',
            'User-Agent': 'okhttp/4.12.0',
            # 'Cookie': 'auth_visit=1; _ga_14VPSGD0HN=GS1.1.1719835902.1.1.1719836588.0.0.0; _ymab_param=w6rHAKlXzdvSgU0nf2aRFae9LFByZ79GB_7LmXUPzYuuQQn2hkp2IbJc176NcfqmVNVUmtmfw0AfGizvhVpaa5qKwJM; tmr_detect=0%7C1719836202990; _ga_Z6YM1FRK5D=GS1.2.1719835904.1.1.1719836379.60.0.0; _dc_gtm_UA-224067-1=1; tmr_lvid=089160966f073ff8ad61ae20d3cc4c53; _ym_visorc=w; sid=7a93fa4ba51e45b1a4584178597c43d5; tmr_lvidTS=1719835903012; atiapp=1; startpage=idatisulogin; _ym_uid=1719835903778768010; uicult=ru-RU; X-Ati-Android-Client=234; AccessToken=293bb1743aef4b9cb4eba530ca69f76f; _ym_isad=2; lastpage=idatisulogin; uicult2=ru; _ym_d=1719835903; _gid=GA1.2.59457443.1719835903; domain_sid=XB1DyqkBgV6bWe41fxizf%3A1719835904377; _ga=GA1.2.46640107.1719835903; _gcl_au=1.1.849499656.1719835901; did=mnBMhODY96nusg3m1X0fVqUHhPNGfbeAWah%2BlfJ4sBk%3D',
        }

        session = requests.Session()

        session.headers.update(headers)
        session.cookies.update(cookies)

        headers = {
            'Content-Type': 'application/json',
        }

        json_data = {
            'country_id': 10,
            'prefix': self.punkt_a,
            'suggestion_types': 23,
        }

        response = requests.post('https://api.ati.su/gw/gis-dict/public/v1/autocomplete/suggestions', headers=headers,
                                 json=json_data)

        punkt_a = response.json().get('suggestions', [{}])[0].get('country', {}).get('id', '')

        json_data = {
            'country_id': 10,
            'prefix': self.punkt_b,
            'suggestion_types': 23,
        }

        response = requests.post('https://api.ati.su/gw/gis-dict/public/v1/autocomplete/suggestions', headers=headers,
                                 json=json_data)

        punkt_b = response.json().get('suggestions', [{}])[0].get('country', {}).get('id', '')

        json_data = {
            "filter": {
                "dates": {
                    "date_option": "today-plus"
                },
                "extra_params": {
                    "adr": False,
                    "adr_type": 0,
                    "conics": False,
                    "do_not_show_regular": False,
                    "dogruz_type": 0,
                    "hydrolysis": False
                },
                "from": {
                    "exact_only": True,
                    "id": punkt_a,
                    "list_id": "",
                    "radius": 0,
                    "type": 0
                },
                "loading_type": 0,
                "to": {
                    "exact_only": True,
                    "id": punkt_b,
                    "list_id": "",
                    "radius": 0,
                    "type": 0
                },
                "truck_type": self.vehicle_type,
                "type_filter": 0,
                "with_rate": False
            },
            "items_per_page": 10,
            "page": 1
        }



        print(json_data)

        response = session.post(
            'https://api.ati.su/mobile/v1.0/trucks/search',
            cookies=cookies,
            headers=headers,
            json=json_data,
        )
        print(response)
        resp = response.json().get('total_count', 0)
        print(resp)
        flattens = []
        if resp:
            for page in range(1, int(int(int(resp) / 100) + 2)):
                json_data['page'] = page
                response = session.post(
                    'https://api.ati.su/mobile/v1.0/trucks/search',
                    cookies=cookies,
                    headers=headers,
                    json=json_data,
                )
                print(response)
                resp = response.json()

                loads = resp.get('trucks', [])
                for load in loads:
                    gorod_a = str(load.get('loading', {}).get('city_id', 9434))
                    gorod_b = str(load.get('unloading', {}).get('city_id', 3058))
                    print(gorod_a, gorod_b)
                    cities = resp.get('cities', {})
                    name_gorod_a = cities.get(gorod_a, {}).get('full_name', '')
                    name_gorod_b = cities.get(gorod_b, {}).get('full_name', '')

                    countries = load.get('route', {}).get('country', 'RUS-KAZ')
                    country_a = countries.split('-')[0]
                    country_b = countries.split('-')[-1]

                    first_date = load.get('first_date', '')
                    last_date = load.get('last_date', '')

                    if first_date:
                        first_date = self.convert_date_to_short_format(first_date)
                    if last_date:
                        last_date = self.convert_date_to_short_format(last_date)

                    transport_number = load.get('transport', {}).get('car_type_bit_mask', 1)
                    type_tran = ' '.join(self.find_combination(transport_number, transport_type))

                    weight = load.get('cargo', {}).get('weight', '')

                    price = load.get('payment', {}).get('sum_without_nds', 'Неизвестная цена')

                    ati_id = load.get('ati_id', 1437455)
                    contact_id1 = load.get('contact_id1', '')
                    contact_id2 = load.get('contact_id2', '')
                    contact_id3 = load.get('contact_id3', '')
                    contact_id4 = load.get('contact_id4', '')

                    acc = resp.get('accounts', {}).get(ati_id, {}).get('contacts', [])
                    telephones = []

                    email = ''

                    for ac in acc:
                        if ac.get('id', 0) in [contact_id1, contact_id2, contact_id3, contact_id4]:

                            telephones.append(ac.get('telephone', '') if ac.get('telephone', '') else '')
                            telephones.append(ac.get('mobile', '') if ac.get('mobile', '') else '')
                            id_ac = ac.get('id', 0)
                            try:
                                response = session.get(
                                    f'https://api.ati.su/mobile/v1.2/firms/contact/{ati_id}/{id_ac}/email'
                                )
                                email = response.json().get('email', '')
                            except:
                                print(
                                    f'Не удалось получить емеил https://api.ati.su/mobile/v1.2/firms/contact/{ati_id}/{id_ac}/email')
                                email = ''

                    telephones = ';'.join(telephones)

                    flattens.append(
                        [name_gorod_a, country_a, name_gorod_b, country_b, f'{first_date}-{last_date}', type_tran,
                         weight, price, f'{telephones} {email}', 'ati-su'])

        column_xlsx = ["Город отправления",
                       "Страна отправления",
                       "Город прибытия",
                       "Страна прибытия", 'Дата', 'Тип машины/груза', 'Вес', 'Цена', 'Контакты', 'Сайт']

        file_path_write = 'output.xlsx'
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Data"
        worksheet.append(column_xlsx)
        workbook.save(file_path_write)

        workbook = load_workbook(file_path_write)
        worksheet = workbook.active
        for rrik in flattens:
            worksheet.append(rrik)
        workbook.save(file_path_write)

