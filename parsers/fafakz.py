
from bs4 import BeautifulSoup
from curl_cffi import requests
from openpyxl import Workbook, load_workbook
import re

session = requests.Session()

headers = {
    'Host': 'm.fa-fa.ru',
    'Accept': 'text/html, */*; q=0.01',
    'User-Agent': 'Mozilla/5.0 (Linux; Android 9; SM-S906N Build/PQ3A.190605.04081832; wv) AppleWebKit/537.36 (KHTML, like Gecko) Version/4.0 Chrome/91.0.4472.114 Mobile Safari/537.36',
    'X-Requested-With': 'XMLHttpRequest',
    'Sec-Fetch-Site': 'same-origin',
    'Sec-Fetch-Mode': 'cors',
    'Sec-Fetch-Dest': 'empty',
    'Referer': 'https://m.fa-fa.ru/',
    # 'Accept-Encoding': 'gzip, deflate',
    'Accept-Language': 'ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7',
    'Connection': 'close',
    # 'Cookie': 'PHPSESSID=s39gour26gore07hlrtogbf4n4; bbuserid=306387; bbpassword=e6b024da4ddf41ecfd8df69d2abd64c9; _ym_uid=1719672596455476887; _ym_d=1719672596; _ymab_param=lucTyg4fB89Nq8j371CmUKczMT-xZMNakFrIEjqFUyjgHhD_GzJvGqwshVdrxf53dKXphenDtdTe1c5jsgFODN4PO9g; _ym_isad=2; _gid=GA1.2.564877948.1719672597; _gat_gtag_UA_113967512_1=1; _ga_295751JDV1=GS1.1.1719672596.1.0.1719672596.60.0.0; _ga=GA1.1.1746340065.1719672597',
}

vehicle_types = {
    "Крытая": -1,
    "Открытая": -2,
    "Любая": 0,
    "Тент": 1,
    "Рефрижератор": 2,
    "Изотерм": 3,
    "Бортовой": 4,
    "Автобус": 5,
    "Автовоз": 6,
    "Автовышка": 7,
    "Автотранспортер": 8,
    "Балковоз": 9,
    "Бензовоз": 10,
    "Бетоновоз": 11,
    "Битумовоз": 12,
    "Газовоз": 13,
    "Зерновоз": 14,
    "Конт.площадка": 15,
    "Кормовоз": 16,
    "Кран": 17,
    "Лесовоз": 18,
    "Манипулятор": 19,
    "Микроавтобус": 20,
    "Муковоз": 21,
    "Открытый Конт.": 22,
    "Панелевоз": 23,
    "Погрузчик": 24,
    "Самосвал": 25,
    "Седельный Тягач": 26,
    "Скотовоз": 27,
    "Трал": 29,
    "Трубовоз": 31,
    "Цельнометалл.": 32,
    "Цементовоз": 33,
    "Цистерна": 34,
    "Шаланда": 35,
    "Щеповоз": 36,
    "Эвакуатор": 37,
    "Экскаватор": 38
}


punkt_a = 'Казахстан'

punkt_b = 'Россия'

def find_vehicle_type(cargo_info):
    for vehicle in vehicle_types:
        if re.search(r'\b' + vehicle[:4], cargo_info, re.IGNORECASE):
            return vehicle
    return "Неизвестный тип"

vehicl_typ = 'Любая'

headers = {
    'Host': 'm.fa-fa.ru',
    # 'Content-Length': '457',
    'Cache-Control': 'max-age=0',
    'Upgrade-Insecure-Requests': '1',
    'Origin': 'https://m.fa-fa.ru',
    'Content-Type': 'application/x-www-form-urlencoded',
    'User-Agent': 'Mozilla/5.0 (Linux; Android 9; SM-S906N Build/PQ3A.190605.04081832; wv) AppleWebKit/537.36 (KHTML, like Gecko) Version/4.0 Chrome/91.0.4472.114 Mobile Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
    'X-Requested-With': 'com.fafa.kz',
    'Sec-Fetch-Site': 'same-origin',
    'Sec-Fetch-Mode': 'navigate',
    'Sec-Fetch-User': '?1',
    'Sec-Fetch-Dest': 'document',
    'Referer': 'https://m.fa-fa.ru/load_m/?sid=32067061',
    # 'Accept-Encoding': 'gzip, deflate',
    'Accept-Language': 'ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7',
    'Connection': 'close',
    # 'Cookie': 'PHPSESSID=s39gour26gore07hlrtogbf4n4; bbuserid=306387; bbpassword=e6b024da4ddf41ecfd8df69d2abd64c9; _ym_uid=1719672596455476887; _ym_d=1719672596; _ymab_param=lucTyg4fB89Nq8j371CmUKczMT-xZMNakFrIEjqFUyjgHhD_GzJvGqwshVdrxf53dKXphenDtdTe1c5jsgFODN4PO9g; _ym_isad=2; _gid=GA1.2.564877948.1719672597; _gat_gtag_UA_113967512_1=1; _ga_295751JDV1=GS1.1.1719672596.1.1.1719673468.13.0.0; _ga=GA1.2.1746340065.1719672597',
}

params = {
    'blank': '1',
}

data = f'City[1]={punkt_a}&Rad[1]=0&Tmidl[2]=0&City[2]=&Rad[2]=0&Tmidl[3]=0&City[3]=&Rad[3]=0&Tmidl[4]=0&City[4]=&Rad[4]=0&Tmidl[5]=0&City[5]=&Rad[5]=0&Tmidl[6]=0&City[6]=&Rad[6]=0&Tmidl[7]=0&City[7]=&Rad[7]=0&Tmidl[8]=0&City[8]=&Rad[8]=0&city_end={punkt_b}&rad_end=0&Ttype={vehicl_typ}&w_ot=0&w_do=22&v_ot=0&v_do=130&dat1=0&dat2=7&type_closed=0&load_search='
response = session.post('https://m.fa-fa.ru/load_m/', params=params,  headers=headers, data=data)
if response.status_code == 302:
    sid = response.headers.get('Location', '').split('sid=')[-1]

soup = BeautifulSoup(response.text, 'html.parser')
options = soup.select('#sid_hist option')

for option in options:
    sid = option.get('value', '-1')
    text = option.text
    if sid != "-1":
        pass



headers = {
    'Host': 'm.fa-fa.kz',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:127.0) Gecko/20100101 Firefox/127.0',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8',
    'Accept-Language': 'ru-RU,ru;q=0.8,en-US;q=0.5,en;q=0.3',
    # 'Accept-Encoding': 'gzip, deflate',
    'Upgrade-Insecure-Requests': '1',
    'Sec-Fetch-Dest': 'document',
    'Sec-Fetch-Mode': 'navigate',
    'Sec-Fetch-Site': 'none',
    'Sec-Fetch-User': '?1',
    'Priority': 'u=1',
    # Requests doesn't support trailers
    # 'Te': 'trailers',
}

params = {
    'sid': sid
}

response = session.get('https://m.fa-fa.kz/load_m/1/', params=params, headers=headers)

print(response, sid)
soup = BeautifulSoup(response.text, 'html.parser')
title_tag = soup.title

# Extract text content from the <title> tag
if title_tag:
    title_text = title_tag.get_text().split('(')[-1].split('.)')[0].strip('�').strip('найдено').strip('шт').strip('')
    if title_text:
        page_end = int(title_text) // 20 + 1

    else:
        page_end = 0

params = {
    'sid': sid
}


flattens = []

pattern = re.compile(r'^res_\d+$')

for page in range(1, page_end+1):
    response = session.get(f'https://m.fa-fa.kz/load_m/{page}/', params=params, headers=headers)
    print(response)
    soup = BeautifulSoup(response.text, 'html.parser')
    for div in soup.find_all('div', id=pattern):
        route = div.find('font', class_='o_dest').text.strip()
        dates = div.find_all('b')[0:2]
        date_from = dates[0].text.strip()
        date_to = dates[1].text.strip()
        date = f"{date_from} - {date_to}"

        # Извлечение информации о типе машины/груза
        cargo_info = div.find_all('td', valign='top')[1].get_text(separator=' ').strip()

        # Поиск типа машины/груза
        cargo_type = find_vehicle_type(cargo_info)

        # Извлечение веса
        weight_match = re.search(r'(\d+)\s*\/', div.find('b', id=re.compile(r'^w_v_\d+$')).text)
        weight = weight_match.group(1) if weight_match else "Неизвестный вес"

        price = div.find_all('tr')[-1].find_all('td')[-1]
        price_info = div.find_all('td', valign='top')[2].get_text(separator=' ').strip()
        price_match = re.search(r'(\d[\d\s\.\,]*\d)', price_info)
        price = price_match.group(0) if price_match else "Неизвестная цена"
        print(price)

        flattens.append([route, date, cargo_type, weight, price, 'fafa-kz'])


column_xlsx = ['Маршрут', 'Дата', 'Тип машины/груза', 'Вес', 'Цена', 'Сайт']
file_path_write = 'output.xlsx'
workbook = Workbook()
worksheet = workbook.active
worksheet.title = "Data"
worksheet.append(column_xlsx)
workbook.save(file_path_write)

workbook = load_workbook(file_path_write)
worksheet = workbook.active
for rrik in flattens:
    worksheet.append(rrik)
workbook.save(file_path_write)






